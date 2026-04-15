#!/usr/bin/env python3
"""
Generate the recommended Modbus register polling spreadsheet for Cat ECS 200.
Creates one sheet per device+port combination, matching the template style from
'CAT ECS 200 SCADA Genset READ v1.1.xlsx'.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy

# ---------------------------------------------------------------------------
# Style definitions (matching the Cat template)
# ---------------------------------------------------------------------------
CAT_YELLOW = "FFFFCD11"
CAT_BLACK = "FF1A1A1A"
CAT_DARK = "FF2C2C2C"
CAT_TEAL = "FF17A2B8"
WHITE = "FFFFFFFF"
LIGHT_GRAY = "FFF0F2F5"
SECTION_BLUE = "FFD6E4F0"
SECTION_GREEN = "FFE2EFDA"
SECTION_AMBER = "FFFFF2CC"
SECTION_RED = "FFFCE4EC"

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

title_font = Font(name="Calibri", size=28, bold=True, color="FFFFFFFF")
title_fill = PatternFill(start_color=CAT_BLACK, end_color=CAT_BLACK, fill_type="solid")

subtitle_font = Font(name="Calibri", size=16, bold=True, color=CAT_BLACK)
subtitle_fill = PatternFill(start_color=CAT_YELLOW, end_color=CAT_YELLOW, fill_type="solid")

header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
header_fill = PatternFill(start_color=CAT_DARK, end_color=CAT_DARK, fill_type="solid")

section_font = Font(name="Calibri", size=12, bold=True, color=CAT_BLACK)

data_font = Font(name="Times New Roman", size=11, bold=False)
data_font_bold = Font(name="Times New Roman", size=11, bold=True)
mono_font = Font(name="Consolas", size=10, bold=False, color="FF555555")

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Column definitions: A-I main, J-M helper
COL_WIDTHS = {
    "A": 52,   # Parameter Name
    "B": 6,    # R/W
    "C": 10,   # Holding Register
    "D": 10,   # Hex Address
    "E": 5,    # Word Count
    "F": 60,   # Description
    "G": 45,   # Scaling
    "H": 20,   # Offset
    "I": 25,   # Limits/Ranges
    "J": 30,   # Block ID
    "K": 10,   # Reg-1
    "L": 10,   # Hex
    "M": 12,   # 0x Hex
}

HEADERS = [
    "Parameter Name",
    "R/W",
    "Holding Register",
    "Hex Addr",
    "Ct",
    "Description",
    "Scaling (Resolution)",
    "Offset",
    "Limits (Ranges)",
    "Block ID",
    "Addr-1",
    "Hex",
    "0x Addr",
]


def hex_addr(reg):
    """Return hex string like 0x009A for a register number."""
    return f"0x{(reg - 1):04X}"


def add_sheet(wb, sheet_name, device_label, port_number, port_note, sections):
    """Create one sheet with the template layout and populate with register data."""
    ws = wb.create_sheet(title=sheet_name)

    # Column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Row 1: Title
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"Cat® ECS 200 — Recommended SCADA Modbus Registers"
    c.font = title_font
    c.fill = title_fill
    c.alignment = center
    ws.row_dimensions[1].height = 40
    for col in range(2, 10):
        cell = ws.cell(row=1, column=col)
        cell.fill = title_fill

    # Row 2: Device / Port
    ws.merge_cells("A2:F2")
    ws.merge_cells("G2:I2")
    c2a = ws["A2"]
    c2a.value = f"DEVICE: {device_label}"
    c2a.font = subtitle_font
    c2a.fill = subtitle_fill
    c2a.alignment = center
    for col in range(2, 7):
        ws.cell(row=2, column=col).fill = subtitle_fill
    c2g = ws["G2"]
    c2g.value = f"PORT {port_number} — {port_note}"
    c2g.font = subtitle_font
    c2g.fill = subtitle_fill
    c2g.alignment = center
    for col in range(8, 10):
        ws.cell(row=2, column=col).fill = subtitle_fill
    ws.row_dimensions[2].height = 30

    # Row 3: Column headers
    for i, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=i)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    ws.row_dimensions[3].height = 35

    # Data rows
    row = 4
    for section in sections:
        # Section header row
        ws.merge_cells(f"A{row}:I{row}")
        cell = ws.cell(row=row, column=1)
        cell.value = section["title"]
        cell.font = section_font
        cell.fill = PatternFill(start_color=section.get("color", SECTION_BLUE),
                                end_color=section.get("color", SECTION_BLUE),
                                fill_type="solid")
        cell.alignment = left_wrap
        cell.border = thin_border
        for col in range(2, 14):
            c = ws.cell(row=row, column=col)
            c.fill = PatternFill(start_color=section.get("color", SECTION_BLUE),
                                 end_color=section.get("color", SECTION_BLUE),
                                 fill_type="solid")
            c.border = thin_border
        ws.row_dimensions[row].height = 24
        row += 1

        # Data rows
        for reg in section["registers"]:
            reg_num = reg["reg"]
            words = reg.get("words", 1)
            name = reg["name"]
            desc = reg.get("desc", "")
            scaling = reg.get("scaling", "")
            offset = reg.get("offset", "")
            limits = reg.get("limits", "")
            block_id = reg.get("block", "")

            ws.cell(row=row, column=1, value=name).font = data_font_bold
            ws.cell(row=row, column=1).alignment = left_wrap
            ws.cell(row=row, column=2, value="R").font = data_font
            ws.cell(row=row, column=2).alignment = center
            ws.cell(row=row, column=3, value=reg_num).font = data_font
            ws.cell(row=row, column=3).alignment = center
            ws.cell(row=row, column=4, value=hex_addr(reg_num)).font = mono_font
            ws.cell(row=row, column=4).alignment = center
            ws.cell(row=row, column=5, value=words).font = data_font
            ws.cell(row=row, column=5).alignment = center
            ws.cell(row=row, column=6, value=desc).font = data_font
            ws.cell(row=row, column=6).alignment = left_top
            ws.cell(row=row, column=7, value=scaling).font = data_font
            ws.cell(row=row, column=7).alignment = left_top
            ws.cell(row=row, column=8, value=offset).font = data_font
            ws.cell(row=row, column=8).alignment = center
            ws.cell(row=row, column=9, value=limits).font = data_font
            ws.cell(row=row, column=9).alignment = left_wrap

            # Block ID
            ws.cell(row=row, column=10, value=block_id).font = Font(name="Consolas", size=10, bold=True, color="FF17A2B8")
            ws.cell(row=row, column=10).alignment = center

            # Helper formulas (K=reg-1, L=hex, M=0x+hex)
            ws.cell(row=row, column=11, value=f"=C{row}-1").font = Font(name="Calibri", size=9)
            ws.cell(row=row, column=12, value=f"=DEC2HEX(K{row})").font = Font(name="Calibri", size=9)
            ws.cell(row=row, column=13, value=f'=CONCAT("0x",TEXT(K{row},"0000"))').font = Font(name="Calibri", size=9)

            # Borders on all cells
            for col in range(1, 14):
                ws.cell(row=row, column=col).border = thin_border

            # Row height based on content
            max_lines = max(
                len(desc.split("\n")) if desc else 1,
                len(scaling.split("\n")) if scaling else 1,
                1,
            )
            ws.row_dimensions[row].height = max(18, min(15 * max_lines, 80))

            # If 2-word register, add the second register on the next line
            if words == 2:
                row += 1
                ws.cell(row=row, column=1, value=f"  ↳ {name} (LSW)").font = Font(name="Times New Roman", size=10, italic=True, color="FF888888")
                ws.cell(row=row, column=1).alignment = left_wrap
                ws.cell(row=row, column=3, value=reg_num + 1).font = Font(name="Times New Roman", size=10, color="FF888888")
                ws.cell(row=row, column=3).alignment = center
                ws.cell(row=row, column=4, value=hex_addr(reg_num + 1)).font = Font(name="Consolas", size=10, color="FF888888")
                ws.cell(row=row, column=4).alignment = center
                ws.cell(row=row, column=6, value="Low word — combine: value = (MSW << 16) | LSW, then apply scaling + offset").font = Font(name="Times New Roman", size=10, italic=True, color="FF888888")
                ws.cell(row=row, column=6).alignment = left_wrap
                ws.cell(row=row, column=10, value=block_id).font = Font(name="Consolas", size=9, color="FFAAAAAA")
                ws.cell(row=row, column=10).alignment = center
                for col in range(1, 14):
                    ws.cell(row=row, column=col).border = thin_border
                ws.row_dimensions[row].height = 18

            row += 1

    # Freeze panes
    ws.freeze_panes = "A4"
    ws.sheet_properties.tabColor = section.get("tab_color", "17A2B8")

    return ws


# ---------------------------------------------------------------------------
# Register data definitions per sheet
# ---------------------------------------------------------------------------

# ===== Sheet 1: ECS 200P — Port 502 — 1s Fast Poll =====
p502_fast = [
    {
        "title": "⚡ BLOCK P1 — Read(2, 4) — Event/Alarm Status",
        "color": SECTION_RED,
        "registers": [
            {"reg": 2, "name": "System Event Lamp Status", "desc": "Indicates system lamp status", "scaling": "Bits 15:4 Unused\nBits 3:2 Amber Lamp (Boolean)\nBits 1:0 Red Lamp (Boolean)", "offset": "", "limits": "Bit-packed", "block": "P1"},
            {"reg": 3, "name": "System Horn Active Status", "desc": "Indicates audible alert horn status", "scaling": "0 = FALSE\n1 = TRUE\n2 = ERROR\n3 = DISABLED OR NOT AVAILABLE", "offset": "", "limits": "0–3", "block": "P1"},
            {"reg": 4, "name": "Hard Shutdown Status", "desc": "Indicates hard shutdown is active", "scaling": "0 = FALSE\n1 = TRUE\n2 = ERROR\n3 = DISABLED OR NOT AVAILABLE", "offset": "", "limits": "0–3", "block": "P1"},
            {"reg": 5, "name": "Soft Shutdown Status", "desc": "Indicates soft shutdown is active", "scaling": "0 = FALSE\n1 = TRUE\n2 = ERROR\n3 = DISABLED OR NOT AVAILABLE", "offset": "", "limits": "0–3", "block": "P1"},
        ],
    },
    {
        "title": "🔋 BLOCK P2 — Read(40, 1) — Battery",
        "color": SECTION_AMBER,
        "registers": [
            {"reg": 40, "name": "Battery Voltage", "desc": "DC battery voltage at controller input", "scaling": "1/20 V / bit", "offset": "0 V", "limits": "0 to 3276.75 V", "block": "P2"},
        ],
    },
    {
        "title": "🚨 BLOCK P3 — Read(101, 10) — User Alarm Groups",
        "color": SECTION_RED,
        "registers": [
            {"reg": 101, "name": "User Alarm Group A", "desc": "Configurable alarm group — 16 user-mapped SPN-FMI events", "scaling": "Bit 0–15 = Events #1–#16\n1 = Alarm ON, 0 = Alarm OFF", "offset": "", "limits": "Bit-packed", "block": "P3"},
            {"reg": 102, "name": "User Alarm Group B", "desc": "Configurable alarm group — 16 user-mapped SPN-FMI events", "scaling": "Bit 0–15 = Events #1–#16", "offset": "", "limits": "Bit-packed", "block": "P3"},
            {"reg": 103, "name": "User Alarm Group C", "desc": "Configurable alarm group — 16 user-mapped SPN-FMI events", "scaling": "Bit 0–15 = Events #1–#16", "offset": "", "limits": "Bit-packed", "block": "P3"},
            {"reg": 104, "name": "User Alarm Group D", "desc": "Configurable alarm group — 16 user-mapped SPN-FMI events", "scaling": "Bit 0–15 = Events #1–#16", "offset": "", "limits": "Bit-packed", "block": "P3"},
            {"reg": 105, "name": "User Alarm Group E", "desc": "Configurable alarm group — 16 user-mapped SPN-FMI events", "scaling": "Bit 0–15 = Events #1–#16", "offset": "", "limits": "Bit-packed", "block": "P3"},
            {"reg": 106, "name": "User Alarm Group F", "desc": "Configurable alarm group — 16 user-mapped SPN-FMI events", "scaling": "Bit 0–15 = Events #1–#16", "offset": "", "limits": "Bit-packed", "block": "P3"},
            {"reg": 107, "name": "User Alarm Group G", "desc": "Configurable alarm group — 16 user-mapped SPN-FMI events", "scaling": "Bit 0–15 = Events #1–#16", "offset": "", "limits": "Bit-packed", "block": "P3"},
            {"reg": 108, "name": "User Alarm Group H", "desc": "Configurable alarm group — 16 user-mapped SPN-FMI events", "scaling": "Bit 0–15 = Events #1–#16", "offset": "", "limits": "Bit-packed", "block": "P3"},
            {"reg": 109, "name": "User Alarm Group I", "desc": "Configurable alarm group — 16 user-mapped SPN-FMI events", "scaling": "Bit 0–15 = Events #1–#16", "offset": "", "limits": "Bit-packed", "block": "P3"},
            {"reg": 110, "name": "User Alarm Group J", "desc": "Configurable alarm group — 16 user-mapped SPN-FMI events", "scaling": "Bit 0–15 = Events #1–#16", "offset": "", "limits": "Bit-packed", "block": "P3"},
        ],
    },
    {
        "title": "⚡ BLOCK P4 — Read(150, 10) — Generator Averages & Power",
        "color": SECTION_BLUE,
        "registers": [
            {"reg": 150, "name": "Generator Average Line-Line Voltage", "desc": "Average Line to Line RMS voltage at generator output", "scaling": "1 V / bit", "offset": "0 V", "limits": "0 to 64255 V", "block": "P4"},
            {"reg": 152, "name": "Generator Average Current", "desc": "Average RMS current at generator output", "scaling": "1 A / bit", "offset": "0 A", "limits": "0 to 64255 A", "block": "P4"},
            {"reg": 154, "name": "Engine Speed", "desc": "Engine RPM (from 200G via CAN)", "scaling": "0.125 rpm / bit", "offset": "0 rpm", "limits": "0 to 8031.875 rpm", "block": "P4"},
            {"reg": 155, "name": "Generator Total Real Power", "words": 2, "desc": "Total real power (kW) delivered by the generator", "scaling": "1 W / bit", "offset": "-2000000000 W", "limits": "-2B to +2.21B W", "block": "P4"},
            {"reg": 157, "name": "Generator Frequency", "desc": "Average AC frequency at generator output", "scaling": "1/128 Hz / bit", "offset": "0 Hz", "limits": "0 to 501.99 Hz", "block": "P4"},
            {"reg": 159, "name": "Generator Average Power Factor", "desc": "Average power factor at generator output", "scaling": "1/16384 / bit", "offset": "-2", "limits": "-2.0 to +2.0", "block": "P4"},
        ],
    },
    {
        "title": "⚡ BLOCK P5 — Read(180, 12) — Per-Phase Voltage/Current/PF",
        "color": SECTION_BLUE,
        "registers": [
            {"reg": 180, "name": "Voltage L1-L2", "desc": "Line to Line RMS voltage phase A-B", "scaling": "1 V / bit", "offset": "0 V", "limits": "0 to 64255 V", "block": "P5"},
            {"reg": 181, "name": "Voltage L2-L3", "desc": "Line to Line RMS voltage phase B-C", "scaling": "1 V / bit", "offset": "0 V", "limits": "0 to 64255 V", "block": "P5"},
            {"reg": 182, "name": "Voltage L3-L1", "desc": "Line to Line RMS voltage phase C-A", "scaling": "1 V / bit", "offset": "0 V", "limits": "0 to 64255 V", "block": "P5"},
            {"reg": 183, "name": "Voltage L1-N", "desc": "Line to Neutral RMS voltage phase A", "scaling": "1 V / bit", "offset": "0 V", "limits": "0 to 64255 V", "block": "P5"},
            {"reg": 184, "name": "Voltage L2-N", "desc": "Line to Neutral RMS voltage phase B", "scaling": "1 V / bit", "offset": "0 V", "limits": "0 to 64255 V", "block": "P5"},
            {"reg": 185, "name": "Voltage L3-N", "desc": "Line to Neutral RMS voltage phase C", "scaling": "1 V / bit", "offset": "0 V", "limits": "0 to 64255 V", "block": "P5"},
            {"reg": 186, "name": "Current L1", "desc": "RMS current phase A", "scaling": "1 A / bit", "offset": "0 A", "limits": "0 to 64255 A", "block": "P5"},
            {"reg": 187, "name": "Current L2", "desc": "RMS current phase B", "scaling": "1 A / bit", "offset": "0 A", "limits": "0 to 64255 A", "block": "P5"},
            {"reg": 188, "name": "Current L3", "desc": "RMS current phase C", "scaling": "1 A / bit", "offset": "0 A", "limits": "0 to 64255 A", "block": "P5"},
            {"reg": 189, "name": "Power Factor L1", "desc": "Power factor phase A", "scaling": "1/16384 / bit", "offset": "-2", "limits": "-2.0 to +2.0", "block": "P5"},
            {"reg": 190, "name": "Power Factor L2", "desc": "Power factor phase B", "scaling": "1/16384 / bit", "offset": "-2", "limits": "-2.0 to +2.0", "block": "P5"},
            {"reg": 191, "name": "Power Factor L3", "desc": "Power factor phase C", "scaling": "1/16384 / bit", "offset": "-2", "limits": "-2.0 to +2.0", "block": "P5"},
        ],
    },
    {
        "title": "🔋 BLOCK P6 — Read(202, 5) — Reactive & Apparent Power",
        "color": SECTION_BLUE,
        "registers": [
            {"reg": 202, "name": "Generator Total Apparent Power", "words": 2, "desc": "Total apparent power (kVA) delivered by the generator", "scaling": "1 VA / bit", "offset": "0 VA", "limits": "0 to 4,294,967,295 VA", "block": "P6"},
            {"reg": 205, "name": "Generator Total Reactive Power", "words": 2, "desc": "Total reactive power (kVAR) delivered by the generator", "scaling": "1 VAr / bit", "offset": "-2000000000 VAr", "limits": "-2B to +2.21B VAr", "block": "P6"},
        ],
    },
    {
        "title": "🔀 BLOCK P7 — Read(228, 5) — Operating State & Breaker",
        "color": SECTION_GREEN,
        "registers": [
            {"reg": 228, "name": "Genset Available Status", "desc": "Whether the genset is available to the paralleling system", "scaling": "0 = Not Available\n1 = Available\n2 = Error\n3 = Disabled Or Not Available", "offset": "", "limits": "0–3", "block": "P7"},
            {"reg": 229, "name": "Generator Circuit Breaker Status", "desc": "Status of the generator circuit breaker", "scaling": "0 = Open\n1 = Closed\n2 = Error\n3 = Disabled Or Not Available", "offset": "", "limits": "0–3", "block": "P7"},
            {"reg": 230, "name": "Operating Mode", "desc": "Current operating mode of the genset controller", "scaling": "0 = Off\n1 = Manual\n2 = Auto\n3 = Test", "offset": "", "limits": "0–3", "block": "P7"},
            {"reg": 231, "name": "Genset Auto Start State", "desc": "Current state of the automatic start/stop sequence", "scaling": "0 = Init\n1 = Pre Crank\n2 = Starting\n3 = Running\n4 = Pre Cooldown\n5 = Cooldown\n6 = Stopping\n7 = Stopped\n8 = Idling", "offset": "", "limits": "0–8", "block": "P7"},
            {"reg": 232, "name": "Engine Run Status", "desc": "Whether the engine is currently running", "scaling": "0 = Off\n1 = Running\n2 = Error\n3 = Disabled Or Not Available", "offset": "", "limits": "0–3", "block": "P7"},
        ],
    },
    {
        "title": "🔀 BLOCK P8 — Read(730, 9) — DBA & Sync Mode",
        "color": SECTION_GREEN,
        "registers": [
            {"reg": 730, "name": "Dead Bus Arbitration (DBA) State", "desc": "Current state of dead bus arbitration logic", "scaling": "Enumerated — see SCADA manual §7.0", "offset": "", "limits": "Enumerated", "block": "P8"},
            {"reg": 731, "name": "DBA Last Fail Reason", "desc": "Reason for last DBA failure", "scaling": "Enumerated failure codes", "offset": "", "limits": "Enumerated", "block": "P8"},
            {"reg": 738, "name": "Synchronize Mode / State", "desc": "Current sync mode and state for paralleling", "scaling": "Enumerated — see SCADA manual §7.0", "offset": "", "limits": "Enumerated", "block": "P8"},
        ],
    },
    {
        "title": "🔗 BLOCK P9 — Read(750, 31+) — System Totals & Per-Unit Data",
        "color": SECTION_GREEN,
        "registers": [
            {"reg": 750, "name": "Number of Units Online", "desc": "Total number of gensets currently online", "scaling": "1 / bit", "offset": "0", "limits": "0 to 32", "block": "P9"},
            {"reg": 751, "name": "Number of Units Available", "desc": "Total number of gensets available", "scaling": "1 / bit", "offset": "0", "limits": "0 to 32", "block": "P9"},
            {"reg": 752, "name": "Number of EPS Units Online", "desc": "Emergency Power Supply units online", "scaling": "1 / bit", "offset": "0", "limits": "0 to 32", "block": "P9"},
            {"reg": 753, "name": "Number of EPS Units Available", "desc": "Emergency Power Supply units available", "scaling": "1 / bit", "offset": "0", "limits": "0 to 32", "block": "P9"},
            {"reg": 756, "name": "System Total Real Power", "words": 2, "desc": "Combined real power from all online gensets", "scaling": "1 W / bit", "offset": "-2000000000 W", "limits": "-2B to +2.21B W", "block": "P9"},
            {"reg": 758, "name": "System Total Percent Load", "desc": "System-wide percent of rated load", "scaling": "0.4 % / bit", "offset": "0 %", "limits": "0 to 100 %", "block": "P9"},
            {"reg": 759, "name": "Gen #1 — Warning Status", "desc": "Gen #1 has active warning", "scaling": "0 = False, 1 = True", "offset": "", "limits": "0–1", "block": "P9"},
            {"reg": 760, "name": "Gen #1 — Shutdown Status", "desc": "Gen #1 has active shutdown", "scaling": "0 = False, 1 = True", "offset": "", "limits": "0–1", "block": "P9"},
            {"reg": 761, "name": "Gen #1 — DBA Status", "desc": "Gen #1 dead bus arbitration status", "scaling": "Enumerated DBA states", "offset": "", "limits": "Enumerated", "block": "P9"},
            {"reg": 762, "name": "Gen #1 — Load Share Status", "desc": "Gen #1 load share state", "scaling": "Enumerated load share states", "offset": "", "limits": "Enumerated", "block": "P9"},
            {"reg": 763, "name": "Gen #1 — Average Voltage", "desc": "Gen #1 average line-line voltage", "scaling": "1 V / bit", "offset": "0 V", "limits": "0 to 64255 V", "block": "P9"},
            {"reg": 764, "name": "Gen #1 — Frequency", "desc": "Gen #1 AC frequency", "scaling": "1/128 Hz / bit", "offset": "0 Hz", "limits": "0 to 501.99 Hz", "block": "P9"},
            {"reg": 765, "name": "Gen #1 — Real Power", "words": 2, "desc": "Gen #1 real power output", "scaling": "1 W / bit", "offset": "-2000000000 W", "limits": "-2B to +2.21B W", "block": "P9"},
            {"reg": 767, "name": "Gen #1 — Breaker Status", "desc": "Gen #1 circuit breaker position", "scaling": "0 = Open\n1 = Closed\n2 = Error\n3 = N/A", "offset": "", "limits": "0–3", "block": "P9"},
            {"reg": 768, "name": "Gen #1 — Operating Mode", "desc": "Gen #1 operating mode", "scaling": "0 = Off\n1 = Manual\n2 = Auto\n3 = Test", "offset": "", "limits": "0–3", "block": "P9"},
            {"reg": 769, "name": "Gen #1 — Genset Available", "desc": "Gen #1 availability", "scaling": "0 = Not Available\n1 = Available", "offset": "", "limits": "0–1", "block": "P9"},
            {"reg": 770, "name": "Gen #1 — Engine Run Status", "desc": "Gen #1 engine running", "scaling": "0 = Off\n1 = Running", "offset": "", "limits": "0–1", "block": "P9"},
            {"reg": 780, "name": "Gen #1 — % Rated kW Load", "desc": "Gen #1 percent of rated kW", "scaling": "0.4 % / bit", "offset": "0 %", "limits": "0 to 100 %", "block": "P9"},
        ],
    },
]

# ===== Sheet 2: ECS 200P — Port 502 — Slow Poll =====
p502_slow = [
    {
        "title": "📈 BLOCK P10 — Read(208, 13) — Energy Meters & Ratings",
        "color": SECTION_BLUE,
        "registers": [
            {"reg": 208, "name": "Positive Real Energy (kWh)", "words": 2, "desc": "Accumulated positive real energy", "scaling": "1 Wh / bit", "offset": "0 Wh", "limits": "0 to 4.29B Wh", "block": "P10"},
            {"reg": 210, "name": "Positive Reactive Energy (kVARh)", "words": 2, "desc": "Accumulated positive reactive energy", "scaling": "1 VARh / bit", "offset": "0 VARh", "limits": "0 to 4.29B VARh", "block": "P10"},
            {"reg": 212, "name": "Negative Real Energy (kWh)", "words": 2, "desc": "Accumulated negative real energy (reverse power)", "scaling": "1 Wh / bit", "offset": "0 Wh", "limits": "0 to 4.29B Wh", "block": "P10"},
            {"reg": 214, "name": "Rated Line-Line Voltage", "desc": "Generator nameplate rated voltage", "scaling": "1 V / bit", "offset": "0 V", "limits": "0 to 64255 V", "block": "P10"},
            {"reg": 215, "name": "Rated Line-Neutral Voltage", "desc": "Generator nameplate rated L-N voltage", "scaling": "1 V / bit", "offset": "0 V", "limits": "0 to 64255 V", "block": "P10"},
            {"reg": 216, "name": "Rated Current", "desc": "Generator nameplate rated current", "scaling": "1 A / bit", "offset": "0 A", "limits": "0 to 64255 A", "block": "P10"},
            {"reg": 217, "name": "Rated Frequency", "desc": "Generator nameplate rated frequency", "scaling": "1/128 Hz / bit", "offset": "0 Hz", "limits": "0 to 501.99 Hz", "block": "P10"},
            {"reg": 218, "name": "Rated Real Power", "words": 2, "desc": "Generator nameplate rated power", "scaling": "1 W / bit", "offset": "-2000000000 W", "limits": "-2B to +2.21B W", "block": "P10"},
            {"reg": 220, "name": "Rated Power Factor", "desc": "Generator nameplate rated power factor", "scaling": "1/16384 / bit", "offset": "-2", "limits": "-2.0 to +2.0", "block": "P10"},
        ],
    },
    {
        "title": "🕐 BLOCK P11 — Read(245, 3) — Real-Time Clock",
        "color": SECTION_AMBER,
        "registers": [
            {"reg": 245, "name": "Date (MM/DD packed)", "desc": "Controller real-time clock date", "scaling": "See SCADA manual §7.12 for byte decode", "offset": "", "limits": "Packed bytes", "block": "P11"},
            {"reg": 246, "name": "Year", "desc": "Controller real-time clock year", "scaling": "See SCADA manual §7.12 for byte decode", "offset": "", "limits": "Packed bytes", "block": "P11"},
            {"reg": 247, "name": "Time (HH:MM:SS packed)", "desc": "Controller real-time clock time", "scaling": "See SCADA manual §7.12 for byte decode", "offset": "", "limits": "Packed bytes", "block": "P11"},
        ],
    },
    {
        "title": "🌐 BLOCK P12 — Read(350, 1) — Data Link Status",
        "color": SECTION_GREEN,
        "registers": [
            {"reg": 350, "name": "Primary Data Link Status", "desc": "CAN datalink health indicator", "scaling": "00 = OK\n01 = Fault\n11 = Disabled Or Not Available", "offset": "", "limits": "0, 1, 3", "block": "P12"},
        ],
    },
    {
        "title": "🌐 BLOCK P13 — Read(577, 6) — Ethernet Port Status",
        "color": SECTION_GREEN,
        "registers": [
            {"reg": 577, "name": "Ethernet 2-Wire Port A Status", "desc": "SPE port A enable status", "scaling": "0 = Disabled\n1 = Enabled", "offset": "", "limits": "0–1", "block": "P13"},
            {"reg": 578, "name": "Ethernet 2-Wire Port B Status", "desc": "SPE port B enable status", "scaling": "0 = Disabled\n1 = Enabled", "offset": "", "limits": "0–1", "block": "P13"},
            {"reg": 579, "name": "Ethernet 4-Wire Port #1 Status", "desc": "Ethernet port #1 enable status", "scaling": "0 = Disabled\n1 = Enabled", "offset": "", "limits": "0–1", "block": "P13"},
            {"reg": 580, "name": "Ethernet 4-Wire Port #2 Status", "desc": "Ethernet port #2 enable status", "scaling": "0 = Disabled\n1 = Enabled", "offset": "", "limits": "0–1", "block": "P13"},
            {"reg": 581, "name": "Ethernet 4-Wire Port #3 Status", "desc": "Ethernet port #3 enable status", "scaling": "0 = Disabled\n1 = Enabled", "offset": "", "limits": "0–1", "block": "P13"},
            {"reg": 582, "name": "Ethernet 4-Wire Port #4 Status", "desc": "Ethernet port #4 enable status", "scaling": "0 = Disabled\n1 = Enabled", "offset": "", "limits": "0–1", "block": "P13"},
        ],
    },
    {
        "title": "🔌 BLOCK P14 — Read(700, 16) — Onboard I/O Status",
        "color": SECTION_AMBER,
        "registers": [
            {"reg": 700, "name": "Onboard Discrete Input Status", "desc": "DI 1–12 status bit-packed", "scaling": "Bit 0–11 = DI1–DI12\n1 = Active, 0 = Inactive", "offset": "", "limits": "Bit-packed", "block": "P14"},
            {"reg": 701, "name": "Onboard Isolated DI Status", "desc": "Isolated digital input status", "scaling": "Bit 0 = Isolated DI", "offset": "", "limits": "Bit-packed", "block": "P14"},
            {"reg": 702, "name": "Onboard Discrete Output Status", "desc": "DO 1–12 status bit-packed", "scaling": "Bit 0–11 = DO1–DO12\n1 = Active, 0 = Inactive", "offset": "", "limits": "Bit-packed", "block": "P14"},
            {"reg": 703, "name": "Onboard Analog Input #1 Value", "desc": "AI #1 sensor value — scaling depends on Cat ET configuration", "scaling": "Sensor-dependent", "offset": "Sensor-dependent", "limits": "Sensor-dependent", "block": "P14"},
            {"reg": 704, "name": "Onboard Analog Input #2 Value", "desc": "AI #2 sensor value", "scaling": "Sensor-dependent", "offset": "Sensor-dependent", "limits": "Sensor-dependent", "block": "P14"},
            {"reg": 705, "name": "Onboard Analog Input #3 Value", "desc": "AI #3 sensor value", "scaling": "Sensor-dependent", "offset": "Sensor-dependent", "limits": "Sensor-dependent", "block": "P14"},
            {"reg": 706, "name": "Onboard Analog Output #1 Value", "words": 2, "desc": "AO #1 output value", "scaling": "Configured output", "offset": "See Cat ET", "limits": "Configured", "block": "P14"},
            {"reg": 708, "name": "Onboard Analog Output #2 Value", "words": 2, "desc": "AO #2 output value", "scaling": "Configured output", "offset": "See Cat ET", "limits": "Configured", "block": "P14"},
            {"reg": 710, "name": "Onboard PWM Output #1 Value", "words": 2, "desc": "PWM #1 output value", "scaling": "Configured output", "offset": "See Cat ET", "limits": "Configured", "block": "P14"},
            {"reg": 712, "name": "Onboard PWM Output #2 Value", "words": 2, "desc": "PWM #2 output value", "scaling": "Configured output", "offset": "See Cat ET", "limits": "Configured", "block": "P14"},
            {"reg": 714, "name": "Onboard PWM Output #3 Value", "words": 2, "desc": "PWM #3 output value", "scaling": "Configured output", "offset": "See Cat ET", "limits": "Configured", "block": "P14"},
        ],
    },
    {
        "title": "⚖️ BLOCK P15 — Read(1465, 5) — LSLD State & Timers",
        "color": SECTION_GREEN,
        "registers": [
            {"reg": 1465, "name": "LSLD Priority Mode", "desc": "Load Sense Load Demand sequencing priority", "scaling": "First On-First Off, Hourmeter, Manual, FIFO-EPS…", "offset": "", "limits": "Enumerated", "block": "P15"},
            {"reg": 1466, "name": "LSLD State", "desc": "Current LSLD state", "scaling": "Idle, Adding, Removing, Load Shed…", "offset": "", "limits": "Enumerated", "block": "P15"},
            {"reg": 1467, "name": "LSLD Desired Units Online", "desc": "Number of units LSLD wants online", "scaling": "1 / bit", "offset": "0", "limits": "0 to 32", "block": "P15"},
            {"reg": 1468, "name": "LSLD Add Timer Remaining", "desc": "Seconds remaining before next unit add", "scaling": "1 s / bit", "offset": "0 s", "limits": "0 to 65535 s", "block": "P15"},
            {"reg": 1469, "name": "LSLD Remove Timer Remaining", "desc": "Seconds remaining before next unit remove", "scaling": "1 s / bit", "offset": "0 s", "limits": "0 to 65535 s", "block": "P15"},
        ],
    },
    {
        "title": "⚖️ BLOCK P16 — Read(1737, 1) — Load Setpoint",
        "color": SECTION_GREEN,
        "registers": [
            {"reg": 1737, "name": "Generator Real Load Control Base Load Setpoint", "desc": "Active base load setpoint as % of rated power", "scaling": "0.01 % / bit", "offset": "-321.27 %", "limits": "-321.27 to 321.28 %", "block": "P16"},
        ],
    },
    {
        "title": "⚖️ BLOCK P17 — Read(1805, 1) — Load Control State",
        "color": SECTION_GREEN,
        "registers": [
            {"reg": 1805, "name": "Generator Real Load Control State", "desc": "Current load control mode", "scaling": "Isochronous, Base Load, Droop, Import/Export…", "offset": "", "limits": "Enumerated", "block": "P17"},
        ],
    },
]

# ===== Sheet 3: ECS 200G — Port 502 — Engine Data =====
g502_engine = [
    {
        "title": "🛢️ BLOCK G1 — Read(41, 9) — Oil/Coolant/Fuel/Intake/Turbo",
        "color": SECTION_AMBER,
        "registers": [
            {"reg": 41, "name": "Oil Pressure", "desc": "Engine oil pressure", "scaling": "1/128 psi / bit", "offset": "0 psi", "limits": "0 to 511 psi", "block": "G1"},
            {"reg": 42, "name": "Coolant Temperature", "desc": "Engine coolant temperature", "scaling": "1/32 °F / bit", "offset": "-273 °F", "limits": "-273 to 1775 °F", "block": "G1"},
            {"reg": 43, "name": "Fuel Level", "desc": "Fuel tank level", "scaling": "1/2.56 % / bit", "offset": "0 %", "limits": "0 to 100 %", "block": "G1"},
            {"reg": 44, "name": "Coolant Level", "desc": "Coolant level sensor (if equipped)", "scaling": "0.5 % / bit", "offset": "0 %", "limits": "0 to 100 %", "block": "G1"},
            {"reg": 45, "name": "Oil Filter Differential Pressure", "desc": "Oil filter dP (if equipped)", "scaling": "1/128 psi / bit", "offset": "0 psi", "limits": "0 to 511 psi", "block": "G1"},
            {"reg": 46, "name": "Intake Manifold Air Temperature", "desc": "Intake manifold air temp", "scaling": "1/32 °F / bit", "offset": "-273 °F", "limits": "-273 to 1775 °F", "block": "G1"},
            {"reg": 49, "name": "Turbocharger Boost Pressure", "desc": "Turbocharger boost pressure", "scaling": "1/128 psi / bit", "offset": "0 psi", "limits": "0 to 511 psi", "block": "G1"},
        ],
    },
    {
        "title": "🌡️ BLOCK G2 — Read(57, 10) — Exhaust/Oil Temp/Fuel Rate/Hours",
        "color": SECTION_AMBER,
        "registers": [
            {"reg": 57, "name": "Exhaust Temperature Left Bank", "desc": "Exhaust manifold temperature — left bank", "scaling": "1/32 °F / bit", "offset": "-273 °F", "limits": "-273 to 1775 °F", "block": "G2"},
            {"reg": 58, "name": "Exhaust Temperature Right Bank", "desc": "Exhaust manifold temperature — right bank", "scaling": "1/32 °F / bit", "offset": "-273 °F", "limits": "-273 to 1775 °F", "block": "G2"},
            {"reg": 59, "name": "Oil Temperature", "desc": "Engine oil temperature", "scaling": "1/32 °F / bit", "offset": "-273 °F", "limits": "-273 to 1775 °F", "block": "G2"},
            {"reg": 62, "name": "Fuel Consumption Rate", "desc": "Instantaneous fuel consumption rate", "scaling": "0.05 L/hr / bit", "offset": "0 L/hr", "limits": "0 to 3212.75 L/hr", "block": "G2"},
            {"reg": 65, "name": "Engine Hours", "words": 2, "desc": "Total engine operating hours", "scaling": "0.05 hr / bit", "offset": "0 hr", "limits": "0 to 210,554,060.75 hr", "block": "G2"},
        ],
    },
    {
        "title": "🔧 BLOCK G3 — Read(154, 1) — Engine Speed",
        "color": SECTION_BLUE,
        "registers": [
            {"reg": 154, "name": "Engine Speed (RPM)", "desc": "Engine crankshaft speed", "scaling": "0.125 rpm / bit", "offset": "0 rpm", "limits": "0 to 8031.875 rpm", "block": "G3"},
        ],
    },
]

# ===== Sheet 4: ECS 200P — Port 50200 — Write Commands =====
p50200_write = [
    {
        "title": "🚦 Alarm Acknowledge & Reset",
        "color": SECTION_RED,
        "registers": [
            {"reg": 1, "name": "Alarm Acknowledge Command", "desc": "Acknowledge active alarms", "scaling": "0 = FALSE\n1 = TRUE", "offset": "", "limits": "0–1", "block": "W-CMD"},
            {"reg": 2, "name": "Reset All Events Command", "desc": "Reset all active events", "scaling": "0 = FALSE\n1 = TRUE", "offset": "", "limits": "0–1", "block": "W-CMD"},
        ],
    },
    {
        "title": "🔀 Mode & Start/Stop Commands",
        "color": SECTION_GREEN,
        "registers": [
            {"reg": 102, "name": "Auto Mode Command", "desc": "Command genset operating mode", "scaling": "Enumerated mode commands", "offset": "", "limits": "Enumerated", "block": "W-CMD"},
            {"reg": 103, "name": "Start Command", "desc": "Engine start command", "scaling": "0 = Off\n1 = Start\n2 = Error\n3 = N/A", "offset": "", "limits": "0–3", "block": "W-CMD"},
            {"reg": 104, "name": "Stop Command", "desc": "Engine stop command", "scaling": "0 = Off\n1 = Stop\n2 = Error\n3 = N/A", "offset": "", "limits": "0–3", "block": "W-CMD"},
        ],
    },
    {
        "title": "⚖️ Setpoint Commands",
        "color": SECTION_BLUE,
        "registers": [
            {"reg": 6, "name": "Base Load Setpoint (kW)", "words": 2, "desc": "Target base load real power setpoint", "scaling": "1 W / bit", "offset": "-2000000000 W", "limits": "-2B to +2.21B W", "block": "W-SET"},
            {"reg": 8, "name": "Base PF Setpoint", "words": 2, "desc": "Target base power factor setpoint", "scaling": "1/16384 / bit", "offset": "-2", "limits": "-2.0 to +2.0", "block": "W-SET"},
            {"reg": 10, "name": "Base VAR Setpoint", "words": 2, "desc": "Target base reactive power setpoint", "scaling": "1 VAr / bit", "offset": "-2000000000 VAr", "limits": "-2B to +2.21B VAr", "block": "W-SET"},
            {"reg": 12, "name": "Base Frequency Setpoint", "words": 2, "desc": "Target frequency bias setpoint", "scaling": "1/128 Hz / bit", "offset": "0 Hz", "limits": "0 to 501.99 Hz", "block": "W-SET"},
            {"reg": 14, "name": "Base Voltage Setpoint", "words": 2, "desc": "Target voltage bias setpoint", "scaling": "1/128 V / bit", "offset": "0 V", "limits": "0 to 511.99 V", "block": "W-SET"},
        ],
    },
]

# ===== Sheet 5: ECS 200P — Port 50201 — Status & Annunciator =====
p50201_status = [
    {
        "title": "🌐 BLOCK S1 — Read(1, 2) — Module Online Status",
        "color": SECTION_GREEN,
        "registers": [
            {"reg": 1, "name": "Paralleling Controller Online", "desc": "Whether the paralleling controller is on the network (always True for self)", "scaling": "0 = Offline\n1 = Online", "offset": "", "limits": "0–1", "block": "S1"},
            {"reg": 2, "name": "Engine ECM Online", "desc": "Whether the engine ECM is communicating via CAN", "scaling": "0 = Offline\n1 = Online", "offset": "", "limits": "0–1", "block": "S1"},
        ],
    },
    {
        "title": "📊 BLOCK S2 — Read(16, 1) — Active Event Counts",
        "color": SECTION_AMBER,
        "registers": [
            {"reg": 16, "name": "Number of Active Events — Genset Controller", "desc": "Count of currently active events on genset controller", "scaling": "1 / bit", "offset": "0", "limits": "0 to 65535", "block": "S2"},
        ],
    },
    {
        "title": "🔔 BLOCK S3 — Read(611, 23) — Annunciator Status Points",
        "color": SECTION_RED,
        "registers": [
            {"reg": 611, "name": "E-Stop Active", "desc": "Engine Auxiliary Switch Status (Emergency Stop) — RED", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 612, "name": "Generator Not In Auto Warning", "desc": "Generator control not in automatic mode — AMBER", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 613, "name": "Overcrank (Failure to Start)", "desc": "Engine failed to start within crank time — RED", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 614, "name": "Overspeed", "desc": "Engine overspeed detected — RED", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 615, "name": "EPS Supplying Load", "desc": "Emergency Power Supply is active — GREEN", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 616, "name": "Overload / Load Shed Contact", "desc": "Overload alarm or load shed active — AMBER", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 617, "name": "High Coolant Temperature (Warning)", "desc": "Coolant temp warning threshold — AMBER", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 618, "name": "High Coolant Temperature (Shutdown)", "desc": "Coolant temp shutdown threshold — RED", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 619, "name": "Low Coolant Temperature Warning", "desc": "Coolant temp low warning — AMBER", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 620, "name": "Low Cranking Voltage Warning", "desc": "Battery voltage too low for cranking — AMBER", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 621, "name": "Air Damper Shutdown", "desc": "Air damper failure shutdown — RED", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 622, "name": "High Battery Voltage (Warning)", "desc": "Battery voltage high warning — AMBER", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 623, "name": "High Battery Voltage (Shutdown)", "desc": "Battery voltage high shutdown — RED", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 624, "name": "Low Battery Voltage Warning", "desc": "Battery voltage low warning — AMBER", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 625, "name": "Low Oil Pressure (Warning)", "desc": "Oil pressure low warning — AMBER", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 626, "name": "Low Oil Pressure (Shutdown)", "desc": "Oil pressure low shutdown — RED", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 627, "name": "Low Coolant Level (Warning)", "desc": "Coolant level low warning — AMBER", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 628, "name": "Low Coolant Level (Shutdown)", "desc": "Coolant level low shutdown — RED", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 629, "name": "Low Fuel Level (Warning)", "desc": "Fuel level low warning — AMBER", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 630, "name": "Low Fuel Level (Shutdown)", "desc": "Fuel level low shutdown — RED", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 631, "name": "Battery Charger AC Failure", "desc": "Battery charger AC supply failure — AMBER", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 632, "name": "SCR SPN 4792 (Warning)", "desc": "SCR aftertreatment warning — AMBER", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
            {"reg": 633, "name": "SCR SPN 4792 (Shutdown)", "desc": "SCR aftertreatment shutdown — RED", "scaling": "0 = FALSE, 1 = TRUE", "offset": "", "limits": "0–1", "block": "S3"},
        ],
    },
]


# ---------------------------------------------------------------------------
# Generate workbook
# ---------------------------------------------------------------------------
def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    add_sheet(wb, "200P — Port 502 Fast",
              "ECS 200 Paralleling (200P)", "502",
              "READ Registers — 1-Second Fast Poll",
              p502_fast)

    add_sheet(wb, "200P — Port 502 Slow",
              "ECS 200 Paralleling (200P)", "502",
              "READ Registers — 5–30 Second Slow Poll",
              p502_slow)

    add_sheet(wb, "200G — Port 502 Engine",
              "ECS 200 Genset (200G)", "502",
              "READ Registers — Engine Sensor Data",
              g502_engine)

    add_sheet(wb, "200P — Port 50200 Write",
              "ECS 200 Paralleling (200P)", "50200",
              "WRITE Registers — Commands & Setpoints (FC16)",
              p50200_write)

    add_sheet(wb, "200P — Port 50201 Status",
              "ECS 200 Paralleling (200P)", "50201",
              "READ Registers — Module Status & Annunciator",
              p50201_status)

    # Change History sheet
    ws_hist = wb.create_sheet(title="Change History")
    ws_hist.column_dimensions["A"].width = 12
    ws_hist.column_dimensions["B"].width = 80
    ws_hist.column_dimensions["C"].width = 30
    for i, h in enumerate(["Version", "Change History", "Notes"], 1):
        c = ws_hist.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border
    ws_hist.cell(row=2, column=1, value="v1.0").border = thin_border
    ws_hist.cell(row=2, column=2, value="Initial recommended register list — curated from LEBE23522-00 SCADA Manual").border = thin_border
    ws_hist.cell(row=2, column=3, value="ECS 200 v1.1 register list").border = thin_border
    ws_hist.sheet_properties.tabColor = "888888"

    out = "ECS 200 Recommended SCADA Registers.xlsx"
    wb.save(out)
    print(f"Generated: {out}")
    print(f"Sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        print(f"  {sn}: {wb[sn].max_row} rows")


if __name__ == "__main__":
    main()
